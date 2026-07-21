#!/usr/bin/env python3
"""
SSH Bulk Login Checker
-----------------------
Tests a single username/password against a list of IP addresses over SSH
and writes the results (Success / Failed / Error + reason) to an Excel file.

Usage:
    python3 ssh_login_checker.py

You will be prompted for:
    - Path to a text file containing one IP (or IP:port) per line
      OR you can just paste IPs interactively
    - Username
    - Password (hidden input)

Output:
    ssh_results.xlsx  -> columns: IP, Port, Status, Detail

Requirements:
    pip install paramiko openpyxl
"""

import getpass
import socket
import sys

import paramiko
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DEFAULT_PORT = 22
CONNECT_TIMEOUT = 8  # seconds


def read_ips_from_file(path):
    ips = []
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                ips.append(line)
    return ips


def read_ips_interactive():
    print("Paste IPs (one per line, or IP:port). Empty line to finish:")
    ips = []
    while True:
        line = input().strip()
        if not line:
            break
        ips.append(line)
    return ips


def parse_target(entry):
    """Return (host, port) from an 'ip' or 'ip:port' string."""
    if ":" in entry:
        host, port_str = entry.rsplit(":", 1)
        try:
            port = int(port_str)
        except ValueError:
            host, port = entry, DEFAULT_PORT
    else:
        host, port = entry, DEFAULT_PORT
    return host, port


def try_ssh_login(host, port, username, password):
    """
    Attempt an SSH login. Returns (status, detail).
    status is one of: "Success", "Failed", "Error"
    """
    client = paramiko.SSHClient()
    # Equivalent of always answering "yes" to the host key fingerprint prompt
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    try:
        client.connect(
            hostname=host,
            port=port,
            username=username,
            password=password,
            timeout=CONNECT_TIMEOUT,
            banner_timeout=CONNECT_TIMEOUT,
            auth_timeout=CONNECT_TIMEOUT,
            allow_agent=False,
            look_for_keys=False,
        )
        # Optional sanity check: run a harmless command to confirm the shell works
        client.exec_command("echo ok", timeout=CONNECT_TIMEOUT)
        return "Success", "Authenticated"
    except paramiko.AuthenticationException:
        return "Failed", "Authentication rejected (bad username/password)"
    except paramiko.SSHException as e:
        return "Error", f"SSH error: {e}"
    except socket.timeout:
        return "Error", "Connection timed out"
    except socket.gaierror:
        return "Error", "Hostname/IP could not be resolved"
    except ConnectionRefusedError:
        return "Error", "Connection refused (port closed or SSH not running)"
    except OSError as e:
        return "Error", f"Network error: {e}"
    except Exception as e:
        return "Error", f"Unexpected error: {e}"
    finally:
        client.close()


def write_excel(results, out_path="ssh_results.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "SSH Results"

    headers = ["IP", "Port", "Status", "Detail"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    status_fill = {
        "Success": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Error": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }

    for host, port, status, detail in results:
        ws.append([host, port, status, detail])
        row_fill = status_fill.get(status)
        if row_fill:
            ws.cell(row=ws.max_row, column=3).fill = row_fill

    # Auto-width columns (rough estimate)
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = length + 4

    wb.save(out_path)
    print(f"\nResults written to: {out_path}")


def main():
    print("=== SSH Bulk Login Checker ===\n")

    choice = input("Load IPs from a file? (y/n): ").strip().lower()
    if choice == "y":
        path = input("Path to IP list file: ").strip()
        try:
            targets = read_ips_from_file(path)
        except FileNotFoundError:
            print(f"File not found: {path}")
            sys.exit(1)
    else:
        targets = read_ips_interactive()

    if not targets:
        print("No IPs provided. Exiting.")
        sys.exit(1)

    username = input("Username: ").strip()
    password = getpass.getpass("Password: ")

    results = []
    total = len(targets)
    for i, entry in enumerate(targets, start=1):
        host, port = parse_target(entry)
        print(f"[{i}/{total}] Trying {host}:{port} ... ", end="", flush=True)
        status, detail = try_ssh_login(host, port, username, password)
        print(f"{status} ({detail})")
        results.append((host, port, status, detail))

    write_excel(results)


if __name__ == "__main__":
    main()
