#!/usr/bin/env python3
"""
SSH Bulk Login Checker - GUI version
--------------------------------------
A simple Tkinter GUI to test one username/password against a list of IPs
over SSH, and export the results to an Excel file.

Requirements:
    pip install paramiko openpyxl

Run:
    python3 ssh_login_checker_gui.py
"""

import socket
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import paramiko
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DEFAULT_PORT = 22
CONNECT_TIMEOUT = 8


# ---------------------------------------------------------------------
# Core SSH logic (same behavior as the CLI version)
# ---------------------------------------------------------------------

def parse_target(entry):
    entry = entry.strip()
    if ":" in entry:
        host, port_str = entry.rsplit(":", 1)
        try:
            port = int(port_str)
        except ValueError:
            host, port = entry, DEFAULT_PORT
    else:
        host, port = entry, DEFAULT_PORT
    return host, port


def try_ssh_login(host, port, username, password):
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        client.connect(
            hostname=host,
            port=port,
            username=username,
            password=password,
            timeout=CONNECT_TIMEOUT,
            banner_timeout=CONNECT_TIMEOUT,
            auth_timeout=CONNECT_TIMEOUT,
            allow_agent=False,
            look_for_keys=False,
        )
        client.exec_command("echo ok", timeout=CONNECT_TIMEOUT)
        return "Success", "Authenticated"
    except paramiko.AuthenticationException:
        return "Failed", "Authentication rejected (bad username/password)"
    except paramiko.SSHException as e:
        return "Error", f"SSH error: {e}"
    except socket.timeout:
        return "Error", "Connection timed out"
    except socket.gaierror:
        return "Error", "Hostname/IP could not be resolved"
    except ConnectionRefusedError:
        return "Error", "Connection refused (port closed or SSH not running)"
    except OSError as e:
        return "Error", f"Network error: {e}"
    except Exception as e:
        return "Error", f"Unexpected error: {e}"
    finally:
        client.close()


def write_excel(results, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SSH Results"

    headers = ["IP", "Port", "Status", "Detail"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    status_fill = {
        "Success": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Error": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }

    for host, port, status, detail in results:
        ws.append([host, port, status, detail])
        fill = status_fill.get(status)
        if fill:
            ws.cell(row=ws.max_row, column=3).fill = fill

    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = length + 4

    wb.save(out_path)


# ---------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------

class SSHCheckerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SSH Bulk Login Checker")
        self.geometry("820x600")
        self.minsize(700, 500)

        self.result_queue = queue.Queue()
        self.results = []
        self.worker_thread = None
        self.stop_requested = False

        self._build_ui()

    # ---------------- UI layout ----------------
    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}

        top_frame = ttk.Frame(self)
        top_frame.pack(fill="x", **pad)

        # Username / Password
        cred_frame = ttk.LabelFrame(top_frame, text="Credentials")
        cred_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(cred_frame, text="Username:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
        self.username_var = tk.StringVar()
        ttk.Entry(cred_frame, textvariable=self.username_var, width=30).grid(
            row=0, column=1, sticky="w", padx=6, pady=4
        )

        ttk.Label(cred_frame, text="Password:").grid(row=0, column=2, sticky="w", padx=6, pady=4)
        self.password_var = tk.StringVar()
        self.password_entry = ttk.Entry(cred_frame, textvariable=self.password_var, width=30, show="*")
        self.password_entry.grid(row=0, column=3, sticky="w", padx=6, pady=4)

        self.show_pw_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            cred_frame, text="Show", variable=self.show_pw_var, command=self._toggle_password
        ).grid(row=0, column=4, sticky="w", padx=6)

        # IP list
        ip_frame = ttk.LabelFrame(top_frame, text="Targets (one IP or IP:port per line)")
        ip_frame.pack(fill="both", expand=False, pady=(0, 8))

        btn_row = ttk.Frame(ip_frame)
        btn_row.pack(fill="x", padx=6, pady=(6, 0))
        ttk.Button(btn_row, text="Load from file...", command=self._load_ip_file).pack(side="left")
        ttk.Button(btn_row, text="Clear", command=self._clear_ips).pack(side="left", padx=6)

        self.ip_text = tk.Text(ip_frame, height=6)
        self.ip_text.pack(fill="both", expand=True, padx=6, pady=6)

        # Controls
        ctrl_frame = ttk.Frame(top_frame)
        ctrl_frame.pack(fill="x")

        self.start_btn = ttk.Button(ctrl_frame, text="Start", command=self._start_scan)
        self.start_btn.pack(side="left")

        self.stop_btn = ttk.Button(ctrl_frame, text="Stop", command=self._stop_scan, state="disabled")
        self.stop_btn.pack(side="left", padx=6)

        self.export_btn = ttk.Button(ctrl_frame, text="Export to Excel...", command=self._export_excel, state="disabled")
        self.export_btn.pack(side="left", padx=6)

        self.progress = ttk.Progressbar(ctrl_frame, mode="determinate")
        self.progress.pack(side="left", fill="x", expand=True, padx=10)

        self.status_label = ttk.Label(ctrl_frame, text="Idle")
        self.status_label.pack(side="right")

        # Results table
        table_frame = ttk.LabelFrame(self, text="Results")
        table_frame.pack(fill="both", expand=True, padx=8, pady=6)

        columns = ("ip", "port", "status", "detail")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        for col, label, width in [
            ("ip", "IP", 160),
            ("port", "Port", 60),
            ("status", "Status", 90),
            ("detail", "Detail", 400),
        ]:
            self.tree.heading(col, text=label)
            self.tree.column(col, width=width, anchor="w")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.tag_configure("Success", background="#c6efce")
        self.tree.tag_configure("Failed", background="#ffc7ce")
        self.tree.tag_configure("Error", background="#ffeb9c")

    # ---------------- UI actions ----------------
    def _toggle_password(self):
        self.password_entry.config(show="" if self.show_pw_var.get() else "*")

    def _load_ip_file(self):
        path = filedialog.askopenfilename(title="Select IP list file", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, "r") as f:
                lines = [line.strip() for line in f if line.strip() and not line.startswith("#")]
            self.ip_text.delete("1.0", "end")
            self.ip_text.insert("1.0", "\n".join(lines))
        except Exception as e:
            messagebox.showerror("Error", f"Could not read file:\n{e}")

    def _clear_ips(self):
        self.ip_text.delete("1.0", "end")

    def _get_targets(self):
        raw = self.ip_text.get("1.0", "end").strip()
        return [line.strip() for line in raw.splitlines() if line.strip()]

    def _start_scan(self):
        targets = self._get_targets()
        username = self.username_var.get().strip()
        password = self.password_var.get()

        if not targets:
            messagebox.showwarning("Missing input", "Please provide at least one IP address.")
            return
        if not username:
            messagebox.showwarning("Missing input", "Please enter a username.")
            return

        # Reset state
        self.results = []
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.progress["value"] = 0
        self.progress["maximum"] = len(targets)
        self.export_btn.config(state="disabled")
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.stop_requested = False

        self.worker_thread = threading.Thread(
            target=self._scan_worker, args=(targets, username, password), daemon=True
        )
        self.worker_thread.start()
        self.after(100, self._poll_queue)

    def _stop_scan(self):
        self.stop_requested = True
        self.status_label.config(text="Stopping...")

    def _scan_worker(self, targets, username, password):
        total = len(targets)
        for i, entry in enumerate(targets, start=1):
            if self.stop_requested:
                self.result_queue.put(("__stopped__", None))
                return
            host, port = parse_target(entry)
            status, detail = try_ssh_login(host, port, username, password)
            self.result_queue.put(("row", (host, port, status, detail, i, total)))
        self.result_queue.put(("__done__", None))

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.result_queue.get_nowait()
                if kind == "row":
                    host, port, status, detail, i, total = payload
                    self.results.append((host, port, status, detail))
                    self.tree.insert("", "end", values=(host, port, status, detail), tags=(status,))
                    self.progress["value"] = i
                    self.status_label.config(text=f"{i}/{total} checked")
                elif kind == "__done__":
                    self._scan_finished("Done")
                    return
                elif kind == "__stopped__":
                    self._scan_finished("Stopped")
                    return
        except queue.Empty:
            pass
        if self.worker_thread and self.worker_thread.is_alive():
            self.after(100, self._poll_queue)

    def _scan_finished(self, status_text):
        self.status_label.config(text=status_text)
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        if self.results:
            self.export_btn.config(state="normal")

    def _export_excel(self):
        if not self.results:
            messagebox.showinfo("No data", "No results to export yet.")
            return
        path = filedialog.asksaveasfilename(
            title="Save results as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="ssh_results.xlsx",
        )
        if not path:
            return
        try:
            write_excel(self.results, path)
            messagebox.showinfo("Saved", f"Results exported to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save file:\n{e}")


if __name__ == "__main__":
    app = SSHCheckerApp()
    app.mainloop()
